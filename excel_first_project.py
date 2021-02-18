import openpyxl
from openpyxl import utils



test_data_wb = openpyxl.load_workbook(filename='data.xlsx')
test_data_ws = test_data_wb.active

filepath = "C:\\dev\\expense-py-project\\test.xlsx"
wb = openpyxl.Workbook()
ws = wb.active

needed_columns = ["Transaction date", "Transaction amount", "Details", "Actual balance"]


def get_column(column_name):
    for cell in test_data_ws[1]:
        if cell.value == column_name:
            # print(cell.column)
            return test_data_ws[openpyxl.utils.get_column_letter(cell.column)]


def print_column(column):
    for cell in column:
        pass
        # print(cell.value)


for i in range(1, len(needed_columns)+1):
    print(i)
    column_buffer = get_column(needed_columns[i-1])
    for idx, cell in enumerate(column_buffer, start=1):
        ws.cell(row=idx, column=i, value=cell.value)
        print(cell.value)

wb.save(filepath)


# for cell in test_data_ws["Transaction amount"]:
#     print(cell.value)