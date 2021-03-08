import openpyxl
from openpyxl import utils
import re
import os
from datetime import datetime
import pandas
from collections import namedtuple
import numpy as np
import shutil
import numbers

output_folder = "output"
combined_output_xlsx = "Combined_output.xlsx"
peti_output_xlsx = "Peti_output"
vanda_output_xlsx = "Vanda_output"
combined_output_path = os.path.join(output_folder, combined_output_xlsx)
peti_output_path = os.path.join(output_folder, peti_output_xlsx)
vanda_output_path = os.path.join(output_folder, vanda_output_xlsx)

config_path = "config.xlsx"

Chart_config = namedtuple('File_format', ['category_col_name', 'category_value_col_name', 'start_cell', 'chart_type',
                                          'scale_X', 'scale_Y'])
str_exp_regex = "Expense regex"
str_exp_amount = "Expense amount"
str_exp_cat = "Expense category"
str_exp_cat_amount = "Expense category amount"
str_exp_o_cat = "Expense other category"
str_exp_o_cat_amount = "Expense other category amount"
str_exp_nature = "Expense nature"
str_exp_nature_amount = "Expense nature amount"
str_information = "Information"
str_information_value = "Information value"
str_exp_description = "Expense description"
str_status = "Status"

expense_category_ntpl = Chart_config(str_exp_cat, str_exp_cat_amount, "M2", 'column', 3, 1)
expense_other_category_ntpl = Chart_config(str_exp_o_cat, str_exp_o_cat_amount, "M17", 'pie', 1, 1)
expense_nature_ntpl = Chart_config(str_exp_nature, str_exp_nature_amount, "M32", 'pie', 1, 1)
information_ntpl = Chart_config(str_information, str_information_value, "U17", 'pie', 1, 1)


source_folder_path = "source_data\\"
output_folder_path = "output\\"

combined_file_name = "Combined_output.xlsx"

File_format = namedtuple('File_format', ['file_name', 'columns', 'header_index', 'source_folder', 'columns_to_merge'])

peti = File_format("Peti_output.xlsx", ["Transaction date", "Transaction amount", "Merged_columns"], 3, source_folder_path + "Peti\\",
                   ["Partner name/Secondary account identifier type", "Transaction type", "Details"])
vanda = File_format("Vanda_output.xlsx", ["Tranzakció időpontja", "Összeg", "Merged_columns"], 13, source_folder_path + "Vanda\\",
                    ["Forgalom típusa", "Ellenoldali név", "Közlemény"])


def get_column(column_name, worksheet):
    for cell in worksheet[1]:
        if cell.value == column_name:
            # print(cell.column)
            return worksheet[openpyxl.utils.get_column_letter(cell.column)]
    raise Exception("Can not find the given column!")


def get_cell_value(row, column, worksheet):
    value = worksheet.cell(row=row, column=column).value
    if isinstance(value, str):
        if value == "None":
            return ""
        else:
            return str(value)
    elif isinstance(value, numbers.Number):
        return int(value)


def get_regex_str(idx, dataframe):
    regex_string = ".*" + dataframe.at[idx, str_exp_regex] + ".*"
    print("regex string: " + regex_string)
    return regex_string


def get_value_list(file_path, column):
    df = pandas.read_excel(file_path, index_col=0)

    value_list = []

    for idx, row in df.iterrows():
        value = df.at[idx, column]
        if not value_list.count(value):
            value_list.append(value)
    if value_list.count(np.nan):
        value_list.remove(np.nan)
    return value_list


def get_and_write_category_amount(exp_category, exp_category_amount, file_path, dataframe_results, dataframe):
    exp_cat_list = get_value_list(file_path, exp_category)

    for idx1, value in enumerate(exp_cat_list):
        expense_in_Ft = 0

        for idx2, row in dataframe.iterrows():
            if dataframe.at[idx2, exp_category] == value:
                expense_in_Ft += dataframe.at[idx2, "Transaction amount"]

        dataframe_results.at[idx1, exp_category] = value
        dataframe_results.at[idx1, exp_category_amount] = expense_in_Ft*(-1)

    return dataframe_results


def write_chart(file_path, chart_config):

    df_res = pandas.read_excel(file_path, index_col=0)

    writer = pandas.ExcelWriter(file_path, engine='xlsxwriter')
    df_res.to_excel(writer, sheet_name='Sheet1')

    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    for chart_idx in range(len(chart_config)):
        cat = df_res[chart_config[chart_idx].category_col_name]
        val = df_res[chart_config[chart_idx].category_value_col_name]

        cat_first_col = df_res.columns.get_loc(chart_config[chart_idx].category_col_name)+1
        cat_first_row = cat.index.get_loc(cat.first_valid_index())+1
        cat_last_col = df_res.columns.get_loc(chart_config[chart_idx].category_col_name)+1
        cat_last_row = cat.index.get_loc(cat.last_valid_index())+1

        val_first_col = df_res.columns.get_loc(chart_config[chart_idx].category_value_col_name)+1
        val_first_row = val.index.get_loc(val.first_valid_index())+1
        val_last_col = df_res.columns.get_loc(chart_config[chart_idx].category_value_col_name)+1
        val_last_row = val.index.get_loc(val.last_valid_index())+1

        pie_chart = (workbook.add_chart({'type': chart_config[chart_idx].chart_type}))

        pie_chart.add_series({
            'name': chart_config[chart_idx].category_col_name,
            'categories': ['Sheet1', cat_first_row, cat_first_col, cat_last_row, cat_last_col],
            'values': ['Sheet1', val_first_row, val_first_col, val_last_row, val_last_col],
            'data_labels': {'value': True, 'percentage': True},
        })

        # Add a title.
        pie_chart.set_title({'name': chart_config[chart_idx].category_col_name})

        # Insert the chart into the worksheet.
        worksheet.insert_chart(chart_config[chart_idx].start_cell, pie_chart,
                               {'x_scale': chart_config[chart_idx].scale_X, 'y_scale': chart_config[chart_idx].scale_Y})

    # Close the Pandas Excel writer and output the Excel file.
    writer.close()


def process_grouped_config_file(df_config):
    print(df_config)

    df_config_column_names = list(df_config.columns)
    df_config_result = pandas.DataFrame(columns=df_config_column_names)

    exp_group_list = []

    for idx1, config in df_config.iterrows():

        if str(df_config.at[idx1, "Expense group"]) == "nan":
            print(str(df_config.at[idx1, "Expense group"]))
            for col in df_config_column_names:
                df_config_result.at[idx1, col] = df_config.at[idx1, col]
        else:
            exp_group = str(df_config.at[idx1, "Expense group"])
            exp_group_list.append(str(df_config.at[idx1, "Expense regex"]))
            exp_regex = ""
            if not exp_group == str(df_config.at[idx1+1, "Expense group"]):
                for idx2 in range(0, len(exp_group_list)):
                    exp_regex += exp_group_list[idx2]
                    if not idx2 == len(exp_group_list)-1:
                        exp_regex = exp_regex + "|"
                exp_regex = "(" + exp_regex + ")"
                for col in df_config_column_names:
                    df_config_result.at[idx1, col] = df_config.at[idx1, col]
                df_config_result.at[idx1, "Expense regex"] = exp_regex
                exp_group_list.clear()

    df_config_result.to_excel("test.xlsx")
    return df_config_result


def generate_expense_category_labels(file_path):
    print()
    print("Generating labels and charts on this file: " + os.path.basename(file_path))
    print(file_path)
    df = pandas.read_excel(file_path, index_col=0)
    df_config = pandas.read_excel(config_path, index_col=0)

    df_config = process_grouped_config_file(df_config)

    df[str_exp_description] = ""
    df[str_exp_cat] = ""
    df[str_exp_o_cat] = ""
    df[str_exp_nature] = ""
    df[str_status] = ""

    df_results = pandas.DataFrame(columns=[str_exp_regex, str_exp_amount,
                                           str_exp_cat, str_exp_cat_amount,
                                           str_exp_o_cat, str_exp_o_cat_amount,
                                           str_exp_nature, str_exp_nature_amount,
                                           str_information, str_information_value])
    num_of_processed_rows = 0
    for idx1, config in df_config.iterrows():
        print("Progress: {}%".format(int((idx1 / len(df_config.index)) * 100)), end='\r')
        expense_in_ft = 0

        regex_string = get_regex_str(idx1, df_config)

        expense_description = df_config.at[idx1, str_exp_description]
        expense_category = df_config.at[idx1, str_exp_cat]
        expense_other_category = df_config.at[idx1, str_exp_o_cat]
        expense_nature = df_config.at[idx1, str_exp_nature]

        for idx2, row2 in df.iterrows():
            if re.search(regex_string, df.at[idx2, "Details"]):
                df.at[idx2, str_exp_description] = expense_description
                df.at[idx2, str_exp_cat] = expense_category
                df.at[idx2, str_exp_o_cat] = expense_other_category
                df.at[idx2, str_exp_nature] = expense_nature
                if not df.at[idx2, str_status] == "progressed":
                    df.at[idx2, str_status] = "progressed"
                    num_of_processed_rows += 1
                else:
                    raise Exception("Multiple found by regex on one row! Need investigation in row " + str(idx2))
                expense_in_ft += df.at[idx2, "Transaction amount"]

        df_results = df_results.append({str_exp_regex: regex_string, str_exp_amount: expense_in_ft*(-1)},
                                       ignore_index=True)

    df_results.at[0, str_information] = "Nem feldolgozott sorok száma"
    df_results.at[1, str_information] = "Feldolgozott sorok száma"
    df_results.at[0, str_information_value] = len(df.index) - num_of_processed_rows
    df_results.at[1, str_information_value] = num_of_processed_rows

    df_results = get_and_write_category_amount(str_exp_cat, str_exp_cat_amount, config_path, df_results, df)
    df_results = get_and_write_category_amount(str_exp_o_cat, str_exp_o_cat_amount, config_path, df_results, df)
    df_results = get_and_write_category_amount(str_exp_nature, str_exp_nature_amount, config_path, df_results, df)

    df.to_excel(file_path)
    result_file_name = os.path.basename(os.path.splitext(file_path)[0]) + "_result.xlsx"
    result_file_path = os.path.join(os.path.dirname(file_path), result_file_name)
    df_results.to_excel(result_file_path)

    print("Expenses labelled succesfully! label rate: {}%".format(int((num_of_processed_rows/len(df.index))*100)))

    chart_config_list = [expense_category_ntpl, expense_other_category_ntpl, expense_nature_ntpl, information_ntpl]
    write_chart(result_file_path, chart_config_list)
    print("Charts inserted")


def get_column_by_name(column_name, max_num_of_lines_from_1, worksheet):
    done_flag = False
    for idx in range(1, max_num_of_lines_from_1):
        row = worksheet[idx]
        for cell in row:
            if cell.value == column_name:
                column = worksheet[openpyxl.utils.get_column_letter(cell.column)]
                done_flag = True
                break
        if done_flag:
            break

    if done_flag:
        return column
    else:
        raise Exception("Can not find the given column!")


def get_column_num_by_name(column_name, max_num_of_lines_from_1, worksheet):
    done_flag = False
    for idx in range(1, max_num_of_lines_from_1):
        row = worksheet[idx]
        for cell in row:
            if cell.value == column_name:
                column = cell.column
                done_flag = True
                break
        if done_flag:
            break

    if done_flag:
        return column
    else:
        raise Exception("Can not find the given column!")


def print_column(column):
    for cell in column:
        print(cell.value)


def create_one_file_form_files(file_format):
    output_file_path = output_folder_path + file_format.file_name

    # Discover files in folder
    file_list = os.listdir(file_format.source_folder)
    files_path_list = []
    for idx, file in enumerate(file_list):
        files_path_list.append(file_format.source_folder + file)

    # Combine all files in the folder in one dataframe
    file_path = files_path_list[0]
    print(file_path)
    df = pandas.read_excel(file_path, header=file_format.header_index, engine="xlrd")
    for idx in range(1, len(files_path_list)):
        print(files_path_list[idx])
        df2 = pandas.read_excel(files_path_list[idx], header=file_format.header_index, engine="xlrd")
        frames = [df, df2]
        df = pandas.concat(frames, ignore_index=True)
        print(idx)

    # Merge the useful columns into one
    df[file_format.columns[2]] = " "
    for col_name in file_format.columns_to_merge:
        df[file_format.columns[2]] += " " + df[col_name].fillna(" ")

    # Keep only the given columns
    df = df[file_format.columns]

    # Remove duplicated rows
    df.drop_duplicates(subset=file_format.columns, keep="first", inplace=True)

    # Rename all columns to unify them
    df.rename(
        columns={df.columns[0]: "Transaction date", df.columns[1]: "Transaction amount", df.columns[2]: "Details"},
        inplace=True)

    # Delete all empty row
    df.dropna(subset=["Transaction date"], inplace=True)

    # Save dataframe into excel file
    df.to_excel(output_file_path)

    # Unify time format
    unify_time_format(output_file_path)

    # Reorder columns by date
    order_rows_by_column(output_file_path, "Transaction date")

    return output_file_path


def order_rows_by_column(file_path, column):
    df = pandas.read_excel(file_path, index_col=0)
    pandas.to_datetime(df[column])
    df = df.sort_values(by=column, ascending=False)
    df = df.reset_index(drop=True)

    df.to_excel(file_path)


def unify_time_format(file_path):
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.active
    column_details = get_column_by_name("Details", 10, ws)
    column_details_num = get_column_num_by_name("Details", 10, ws)
    column_time_num = get_column_num_by_name("Transaction date", 10, ws)

    for idx in range(2, len(column_details) + 1):
        result = re.search(".*([0-9][0-9])([0-1][0-9])([0-3][0-9])([0-2][0-9]):([0-6][0-9]).*",
                           str(ws.cell(row=idx, column=column_details_num).value))
        result2 = re.search(".*([0-9]{4}).([0-1][0-9]).([0-3][0-9]) ([0-2][0-9]):([0-6][0-9]):([0-6][0-9]).*",
                            str(ws.cell(row=idx, column=column_details_num).value))
        result3 = re.search(".*([0-1][0-9])-([0-3][0-9])-([0-9]{4}).*",
                            str(ws.cell(row=idx, column=column_time_num).value))
        result4 = re.search(".*([0-9]{4}).([0-1][0-9]).([0-3][0-9]). ([0-2][0-9]):([0-6][0-9]):([0-6][0-9]).*",
                            str(ws.cell(row=idx, column=column_time_num).value))

        if result:
            date = datetime(int(result.group(1)) + 2000, int(result.group(2)), int(result.group(3)),
                            int(result.group(4)), int(result.group(5)))
            # print(date)
        elif result2:
            date = datetime(int(result2.group(1)), int(result2.group(2)), int(result2.group(3)), int(result2.group(4)),
                            int(result2.group(5)), int(result2.group(6)))
            # print(date)
        elif result3:
            date = datetime(int(result3.group(3)), int(result3.group(1)), int(result3.group(2)))
            # print(date)
        elif result4:
            date = datetime(int(result4.group(1)), int(result4.group(2)), int(result4.group(3)), int(result4.group(4)),
                            int(result4.group(5)), int(result4.group(6)))
            # print(date)
        else:
            print(str(ws.cell(row=idx, column=column_details_num).value))
            print(str(ws.cell(row=idx, column=column_time_num).value))
            print(str(ws.cell(row=idx, column=column_time_num).value))
            raise Exception("Can not detect time!")

        ws.cell(row=idx, column=column_time_num, value=str(date))

    wb.save(file_path)


def combine_two_files(file1, file2):
    output_file_path = output_folder_path + combined_file_name

    df1 = pandas.read_excel(file1, index_col=0)
    df2 = pandas.read_excel(file2, index_col=0)

    frames = [df1, df2]
    df = pandas.concat(frames, ignore_index=True)

    df.to_excel(output_file_path)

    order_rows_by_column(output_file_path, "Transaction date")

    return output_file_path


def separate_file_by_date(file_path, folder_name):
    month_list = get_date_list(file_path, False)
    year_list = get_date_list(file_path, True)

    if not os.path.exists(output_folder_path + folder_name):
        os.makedirs(output_folder_path + folder_name)
    else:
        shutil.rmtree(output_folder_path + folder_name)
        os.makedirs(output_folder_path + folder_name)

    df = pandas.read_excel(file_path, index_col=0)
    for date in month_list:
        df_filter = df["Transaction date"].str.contains(".*" + date + ".*")
        df2 = df[df_filter]
        df2 = df2.reset_index(drop=True)
        result = re.search("([0-9]{4})", date)
        if result:
            if not os.path.exists(output_folder_path + folder_name + "\\" + result.group(0) + " monthly output"):
                os.makedirs(output_folder_path + folder_name + "\\" + result.group(0) + " monthly output")
            df2.to_excel(output_folder_path + folder_name + "\\" + result.group(0) + " monthly output" + "\\" + date + "_output.xlsx")
        else:
            raise Exception("Can not detect year!")

    for date in year_list:
        df_filter = df["Transaction date"].str.contains(".*" + date + ".*")
        df2 = df[df_filter]
        df2 = df2.reset_index(drop=True)
        df2.to_excel(output_folder_path + folder_name + "\\" + date + "_output.xlsx")


def get_date_list(file_path, only_year):
    wb = openpyxl.load_workbook(filename=file_path)
    ws = wb.active
    column_time = get_column_by_name("Transaction date", 10, ws)
    column_time_num = get_column_num_by_name("Transaction date", 10, ws)

    date_list = []

    for idx in range(2, len(column_time) + 1):

        result = re.search("([0-9]{4})-([0-1][0-9])", str(ws.cell(row=idx, column=column_time_num).value))
        if result:
            if only_year:
                date = str(result.group(1))
            else:
                date = str(result.group(0))
            if not date_list.count(date):
                date_list.append(date)
    return date_list


def add_more_columns(folder_path):
    # Discover files in folder
    file_list = os.listdir(folder_path)
    files_path_list = []
    for idx, file in enumerate(file_list):
        files_path_list.append(folder_path + file)


def filter_all_vendor(file_path):
    df = pandas.read_excel(file_path, index_col=0)
    str_details = "Details"
    vendor_list = []
    vendor_pcs_list = []

    for idx1, row in df.iterrows():
        result = re.search(".* [0-9a-zA-Z]{8} (.*) [0-9]{8}:[0-9]{2}", df.at[idx1, str_details])
        if result and (not df.at[idx1, "Status"] == "progressed"):
            if not vendor_list.count(result.group(1)):
                vendor_list.append(result.group(1))
                vendor_pcs_list.append(1)
            else:
                vendor_name_idx = vendor_list.index(result.group(1))
                vendor_pcs_list[vendor_name_idx] = vendor_pcs_list[vendor_name_idx]+1

    df_vendor = pandas.DataFrame(list(zip(vendor_list, vendor_pcs_list)), columns=["Vendor name", "Vendor occurrence"])
    df_vendor.to_excel("test_vendor_names.xlsx")


if not os.path.exists("output"):
    os.makedirs("output")
else:
    shutil.rmtree("output")
    os.makedirs("output")

output_file_path_peti = create_one_file_form_files(peti)
#output_file_path_vanda = create_one_file_form_files(vanda)

#output_combined_file_path = combine_two_files(output_file_path_peti, output_file_path_vanda)

# separate_file_by_date(output_file_path_peti, "Peti separated files")
# separate_file_by_date(output_file_path_vanda, "Vanda separated files")
# separate_file_by_date(output_combined_file_path, "Combined separated files")

# add_more_columns(output_folder_path)

listOfFiles = list()

for (dirpath, dirnames, filenames) in os.walk("output"):
    listOfFiles += [os.path.join(dirpath, file) for file in filenames]

for file in listOfFiles:
    generate_expense_category_labels(file)

filter_all_vendor(output_file_path_peti)


